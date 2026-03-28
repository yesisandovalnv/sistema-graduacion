from django.db.models import Avg, Case, Count, DurationField, ExpressionWrapper, F, FloatField, Max, Q, Value, When
from django.db.models.functions import Coalesce, Now, TruncMonth
from io import BytesIO
import os
import zlib
from django.conf import settings
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from documentos.models import DocumentoPostulacion
from modalidades.models import Modalidad
from postulantes.models import Postulacion, Postulante


def porcentaje_avance_postulacion(postulacion_id: int) -> float:
    queryset = Postulacion.objects.filter(pk=postulacion_id).annotate(
        total_etapas=Count('modalidad__etapas', filter=Q(modalidad__etapas__activo=True), distinct=True)
    )
    queryset = queryset.annotate(
        etapas_completadas=Case(
            When(estado_general='TITULADO', then=F('total_etapas')),
            When(etapa_actual__orden__isnull=False, then=F('etapa_actual__orden') - Value(1)),
            default=Value(0),
        )
    ).annotate(
        porcentaje=Case(
            When(total_etapas=0, then=Value(0.0)),
            default=ExpressionWrapper(
                100.0 * F('etapas_completadas') / F('total_etapas'),
                output_field=FloatField(),
            ),
            output_field=FloatField(),
        )
    )

    result = queryset.values('porcentaje').first()
    if not result:
        return 0.0
    return round(result['porcentaje'] or 0.0, 2)


def documentos_rechazados_por_postulacion(postulacion_id: int) -> dict:
    rejected_docs = list(
        DocumentoPostulacion.objects.filter(postulacion_id=postulacion_id, estado='rechazado')
        .select_related('tipo_documento')
        .values(
            'id',
            'tipo_documento_id',
            'tipo_documento__nombre',
            'comentario_revision',
            'fecha_revision',
        )
        .order_by('-fecha_revision', '-id')
    )
    return {
        'postulacion_id': postulacion_id,
        'total_rechazados': len(rejected_docs),
        'documentos': rejected_docs,
    }


def dashboard_general(fecha_inicio=None, fecha_fin=None, year=None) -> dict:
    """
    Dashboard general con validaciones defensivas para NULL values.
    FASE 3: Incluye métricas calculadas (sin hardcode):
    - tasa_aprobacion
    - promedio_procesamiento_dias
    - satisfaccion_score
    - proyeccion_mes_porcentaje
    """
    from django.utils import timezone
    from dateutil.relativedelta import relativedelta
    
    try:
        # --- SECCIÓN 1: Conteos simples (SAFE) ---
        try:
            total_postulantes = Postulante.objects.count() or 0
            total_postulaciones = Postulacion.objects.count() or 0
            total_modalidades = Modalidad.objects.filter(activa=True).count() or 0
        except Exception as e:
            print(f"⚠️ Error en conteos simples: {e}")
            total_postulantes = 0
            total_postulaciones = 0
            total_modalidades = 0
        
        # --- SECCIÓN 2: Resumen de documentos CONSOLIDADO (SAFE NULL HANDLING) ---
        total_documentos = 0
        docs_pendientes = 0
        docs_rechazados = 0
        try:
            documentos_resumen = DocumentoPostulacion.objects.aggregate(
                total_documentos=Count('id'),
                documentos_pendientes=Count('id', filter=Q(estado='pendiente')),
                documentos_rechazados=Count('id', filter=Q(estado='rechazado')),
                documentos_aprobados=Count('id', filter=Q(estado='aprobado')),
            )
            total_documentos = int(documentos_resumen.get('total_documentos') or 0)
            docs_pendientes = int(documentos_resumen.get('documentos_pendientes') or 0)
            docs_rechazados = int(documentos_resumen.get('documentos_rechazados') or 0)
            docs_aprobados = int(documentos_resumen.get('documentos_aprobados') or 0)
        except Exception as e:
            print(f"⚠️ Error en documentos_resumen: {e}")
            total_documentos = 0
            docs_pendientes = 0
            docs_rechazados = 0
            docs_aprobados = 0
        
        # --- SECCIÓN 3: Postulaciones por estado (SAFE NULL HANDLING) ---
        postulaciones_estado_list = []
        try:
            postulaciones_por_estado = list(
                Postulacion.objects.values('estado_general')
                .annotate(total=Count('id'))
                .order_by('-total')
            )
            postulaciones_estado_list = [
                {
                    'estado': (item.get('estado_general') or 'Sin estado'),
                    'total': int(item.get('total') or 0)
                }
                for item in postulaciones_por_estado
            ]
        except Exception as e:
            print(f"⚠️ Error en postulaciones_por_estado: {e}")
            postulaciones_estado_list = []
        
        # --- SECCIÓN 4: Total titulados (SAFE) ---
        total_titulados = 0
        try:
            total_titulados = Postulacion.objects.filter(estado_general='TITULADO').count() or 0
        except Exception as e:
            print(f"⚠️ Error en total_titulados: {e}")
            total_titulados = 0

        # --- SECCIÓN 5: MÉTRICAS CALCULADAS (NUEVAS - FASE 3) ---
        
        # Métrica 1: Tasa de Aprobación
        tasa_aprobacion = 0.0
        try:
            if total_postulaciones > 0:
                tasa_aprobacion = round((total_titulados / total_postulaciones) * 100, 2)
            print(f"✅ Tasa Aprobación: {tasa_aprobacion}% ({total_titulados}/{total_postulaciones})")
        except Exception as e:
            print(f"⚠️ Error calculando tasa_aprobacion: {e}")
            tasa_aprobacion = 0.0
        
        # Métrica 2: Promedio Procesamiento (días)
        promedio_procesamiento_dias = 0.0
        try:
            # Buscar postulaciones completadas (TITULADO o APROBADO)
            postulaciones_completadas = Postulacion.objects.filter(
                estado_general__in=['TITULADO', 'APROBADO']
            ).annotate(
                dias_proceso=ExpressionWrapper(
                    Now() - F('fecha_postulacion'),
                    output_field=DurationField()
                )
            ).aggregate(
                promedio_dias=Avg('dias_proceso')
            )
            
            if postulaciones_completadas.get('promedio_dias'):
                promedio_timedelta = postulaciones_completadas['promedio_dias']
                promedio_procesamiento_dias = round(promedio_timedelta.total_seconds() / 86400, 2)
            print(f"✅ Promedio Procesamiento: {promedio_procesamiento_dias} días")
        except Exception as e:
            print(f"⚠️ Error calculando promedio_procesamiento: {e}")
            promedio_procesamiento_dias = 0.0
        
        # Métrica 3: Satisfacción (basada en ratio documentos aprobados)
        satisfaccion_score = "N/A"
        try:
            if total_documentos > 0:
                # Satisfacción = (docs_aprobados / total_documentos) * 10
                satisfaccion_score = round((docs_aprobados / total_documentos) * 10, 2)
                # Limitar a máximo 10
                satisfaccion_score = min(satisfaccion_score, 10.0)
                print(f"✅ Satisfacción: {satisfaccion_score}/10 ({docs_aprobados}/{total_documentos})")
            else:
                print(f"ℹ️  Satisfacción: N/A (sin documentos para calcular)")
        except Exception as e:
            print(f"⚠️ Error calculando satisfaccion: {e}")
            satisfaccion_score = "N/A"
        
        # Métrica 4: Proyección Mes (comparar mes actual vs mes anterior)
        proyeccion_mes_porcentaje = 0.0
        try:
            ahora = timezone.now()
            mes_actual_inicio = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            mes_anterior_inicio = (mes_actual_inicio - relativedelta(months=1))
            mes_anterior_fin = mes_actual_inicio
            
            postulaciones_mes_anterior = Postulacion.objects.filter(
                fecha_postulacion__gte=mes_anterior_inicio,
                fecha_postulacion__lt=mes_actual_inicio
            ).count()
            
            postulaciones_mes_actual = Postulacion.objects.filter(
                fecha_postulacion__gte=mes_actual_inicio,
                fecha_postulacion__lte=ahora
            ).count()
            
            if postulaciones_mes_anterior > 0:
                proyeccion_mes_porcentaje = round(
                    ((postulaciones_mes_actual - postulaciones_mes_anterior) / postulaciones_mes_anterior) * 100,
                    2
                )
            else:
                # Si no hay datos del mes anterior, usar cambio absoluto
                proyeccion_mes_porcentaje = round((postulaciones_mes_actual / max(1, postulaciones_mes_anterior + 1)) * 100, 2)
            
            print(f"✅ Proyección Mes: {proyeccion_mes_porcentaje}% (act: {postulaciones_mes_actual}, ant: {postulaciones_mes_anterior})")
        except Exception as e:
            print(f"⚠️ Error calculando proyeccion_mes: {e}")
            proyeccion_mes_porcentaje = 0.0
        
        # --- SECCIÓN 6: CAMBIOS MES-A-MES (nuevos - para KPIs) ---
        
        # Cambio Postulantes mes-a-mes
        cambio_postulantes_porcentaje = 0.0
        try:
            ahora = timezone.now()
            mes_actual_inicio = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            mes_anterior_inicio = (mes_actual_inicio - relativedelta(months=1))
            
            postulantes_mes_anterior = Postulante.objects.filter(
                created_at__gte=mes_anterior_inicio,
                created_at__lt=mes_actual_inicio
            ).count() if hasattr(Postulante, 'created_at') else 0
            
            postulantes_mes_actual = Postulante.objects.filter(
                created_at__gte=mes_actual_inicio,
                created_at__lte=ahora
            ).count() if hasattr(Postulante, 'created_at') else 0
            
            if postulantes_mes_anterior > 0:
                cambio_postulantes_porcentaje = round(
                    ((postulantes_mes_actual - postulantes_mes_anterior) / postulantes_mes_anterior) * 100,
                    2
                )
            print(f"✅ Cambio Postulantes: {cambio_postulantes_porcentaje}%")
        except Exception as e:
            print(f"⚠️ Error calculando cambio_postulantes: {e}")
            cambio_postulantes_porcentaje = 0.0
        
        # Cambio Documentos mes-a-mes
        cambio_documentos_porcentaje = 0.0
        try:
            ahora = timezone.now()
            mes_actual_inicio = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            mes_anterior_inicio = (mes_actual_inicio - relativedelta(months=1))
            
            documentos_mes_anterior = DocumentoPostulacion.objects.filter(
                fecha_subida__gte=mes_anterior_inicio,
                fecha_subida__lt=mes_actual_inicio
            ).count()
            
            documentos_mes_actual = DocumentoPostulacion.objects.filter(
                fecha_subida__gte=mes_actual_inicio,
                fecha_subida__lte=ahora
            ).count()
            
            if documentos_mes_anterior > 0:
                cambio_documentos_porcentaje = round(
                    ((documentos_mes_actual - documentos_mes_anterior) / documentos_mes_anterior) * 100,
                    2
                )
            print(f"✅ Cambio Documentos: {cambio_documentos_porcentaje}%")
        except Exception as e:
            print(f"⚠️ Error calculando cambio_documentos: {e}")
            cambio_documentos_porcentaje = 0.0
        
        # Cambio Titulados mes-a-mes
        cambio_titulados_porcentaje = 0.0
        try:
            ahora = timezone.now()
            mes_actual_inicio = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            mes_anterior_inicio = (mes_actual_inicio - relativedelta(months=1))
            
            titulados_mes_anterior = Postulacion.objects.filter(
                fecha_postulacion__gte=mes_anterior_inicio,
                fecha_postulacion__lt=mes_actual_inicio,
                estado_general='TITULADO'
            ).count()
            
            titulados_mes_actual = Postulacion.objects.filter(
                fecha_postulacion__gte=mes_actual_inicio,
                fecha_postulacion__lte=ahora,
                estado_general='TITULADO'
            ).count()
            
            if titulados_mes_anterior > 0:
                cambio_titulados_porcentaje = round(
                    ((titulados_mes_actual - titulados_mes_anterior) / titulados_mes_anterior) * 100,
                    2
                )
            print(f"✅ Cambio Titulados: {cambio_titulados_porcentaje}%")
        except Exception as e:
            print(f"⚠️ Error calculando cambio_titulados: {e}")
            cambio_titulados_porcentaje = 0.0
        
        # Cambio Tasa de Aprobación mes-a-mes
        cambio_tasa_porcentaje = 0.0
        try:
            ahora = timezone.now()
            mes_actual_inicio = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            mes_anterior_inicio = (mes_actual_inicio - relativedelta(months=1))
            
            # Calcular tasa anterior
            total_postulaciones_anterior = Postulacion.objects.filter(
                fecha_postulacion__gte=mes_anterior_inicio,
                fecha_postulacion__lt=mes_actual_inicio
            ).count()
            
            titulados_anterior = Postulacion.objects.filter(
                fecha_postulacion__gte=mes_anterior_inicio,
                fecha_postulacion__lt=mes_actual_inicio,
                estado_general='TITULADO'
            ).count()
            
            tasa_anterior = (titulados_anterior / total_postulaciones_anterior * 100) if total_postulaciones_anterior > 0 else 0
            
            # Calcular tasa actual  
            total_postulaciones_actual = Postulacion.objects.filter(
                fecha_postulacion__gte=mes_actual_inicio,
                fecha_postulacion__lte=ahora
            ).count()
            
            titulados_actual = Postulacion.objects.filter(
                fecha_postulacion__gte=mes_actual_inicio,
                fecha_postulacion__lte=ahora,
                estado_general='TITULADO'
            ).count()
            
            tasa_actual = (titulados_actual / total_postulaciones_actual * 100) if total_postulaciones_actual > 0 else 0
            
            # Cambio en puntos porcentuales
            cambio_tasa_porcentaje = round(tasa_actual - tasa_anterior, 2)
            print(f"✅ Cambio Tasa: {cambio_tasa_porcentaje}% (ant: {tasa_anterior}%, act: {tasa_actual}%)")
        except Exception as e:
            print(f"⚠️ Error calculando cambio_tasa: {e}")
            cambio_tasa_porcentaje = 0.0

        # --- RETORNO: Estructura garantizada (NUNCA NULL) ---
        return {
            'total_postulaciones': int(total_postulaciones),
            'total_postulantes': int(total_postulantes),
            'total_modalidades': int(total_modalidades),
            'total_documentos': int(total_documentos),
            'postulaciones_por_estado_general': postulaciones_estado_list,
            'postulaciones_por_mes': [],
            'documentos_por_mes': [],
            'tiempo_promedio_por_mes': [],
            'postulaciones_por_etapa': [],
            'documentos_pendientes': int(docs_pendientes),
            'documentos_rechazados': int(docs_rechazados),
            'total_titulados': int(total_titulados),
            'tiempo_promedio_proceso_dias': promedio_procesamiento_dias,
            # NUEVOS (FASE 3) - Métricas sin hardcode
            'tasa_aprobacion': tasa_aprobacion,
            'promedio_procesamiento_dias': promedio_procesamiento_dias,
            'satisfaccion_score': satisfaccion_score,  # N/A si sin datos, número si con datos
            'proyeccion_mes_porcentaje': proyeccion_mes_porcentaje,
            # NUEVOS (FASE 4) - Cambios mes-a-mes para KPIs
            'cambio_postulantes_porcentaje': cambio_postulantes_porcentaje,
            'cambio_documentos_porcentaje': cambio_documentos_porcentaje,
            'cambio_titulados_porcentaje': cambio_titulados_porcentaje,
            'cambio_tasa_porcentaje': cambio_tasa_porcentaje,
        }
        
    except Exception as e:
        import traceback
        print(f"❌ ERROR CRÍTICO en dashboard_general: {str(e)}")
        print(traceback.format_exc())
        # Safety net - SIEMPRE devuelve estructura válida
        return {
            'total_postulaciones': 0,
            'total_postulantes': 0,
            'total_modalidades': 0,
            'total_documentos': 0,
            'postulaciones_por_estado_general': [],
            'postulaciones_por_mes': [],
            'documentos_por_mes': [],
            'tiempo_promedio_por_mes': [],
            'postulaciones_por_etapa': [],
            'documentos_pendientes': 0,
            'documentos_rechazados': 0,
            'total_titulados': 0,
            'tiempo_promedio_proceso_dias': 0.0,
            'tasa_aprobacion': 0.0,
            'promedio_procesamiento_dias': 0.0,
            'satisfaccion_score': "N/A",  # N/A por defecto (sin datos)
            'proyeccion_mes_porcentaje': 0.0,
            'cambio_postulantes_porcentaje': 0.0,
            'cambio_documentos_porcentaje': 0.0,
            'cambio_titulados_porcentaje': 0.0,
            'cambio_tasa_porcentaje': 0.0,
        }


def get_dashboard_chart_data(meses: int = 6) -> dict:
    """
    Obtiene datos históricos formateados para los gráficos de Charts.jsx
    Retorna estructura con lineChartData, barChartData, pieChartData
    """
    import logging
    from django.utils import timezone
    from dateutil.relativedelta import relativedelta
    
    logger = logging.getLogger(__name__)
    
    try:
        # Estructura de respuesta con datos mock como respaldo
        mock_response = {
            'lineChartData': [
                {'mes': 'Ene', 'graduados': 45, 'pendientes': 120, 'aprobados': 95},
                {'mes': 'Feb', 'graduados': 72, 'pendientes': 98, 'aprobados': 142},
                {'mes': 'Mar', 'graduados': 98, 'pendientes': 76, 'aprobados': 165},
                {'mes': 'Abr', 'graduados': 125, 'pendientes': 62, 'aprobados': 189},
                {'mes': 'May', 'graduados': 145, 'pendientes': 48, 'aprobados': 210},
                {'mes': 'Jun', 'graduados': 156, 'pendientes': 42, 'aprobados': 248},
            ],
            'barChartData': [
                {'semana': 'Sem 1', 'postulantes': 45, 'documentos': 38},
                {'semana': 'Sem 2', 'postulantes': 52, 'documentos': 48},
                {'semana': 'Sem 3', 'postulantes': 38, 'documentos': 35},
                {'semana': 'Sem 4', 'postulantes': 61, 'documentos': 55},
                {'semana': 'Sem 5', 'postulantes': 58, 'documentos': 52},
                {'semana': 'Sem 6', 'postulantes': 72, 'documentos': 68},
            ],
            'pieChartData': [
                {'name': 'Completado', 'value': 45, 'color': '#10b981'},
                {'name': 'En Proceso', 'value': 30, 'color': '#f59e0b'},
                {'name': 'Por Revisar', 'value': 15, 'color': '#3b82f6'},
                {'name': 'Rechazado', 'value': 10, 'color': '#ef4444'},
            ],
            'error': None
        }
        
        # Validar meses
        if not isinstance(meses, int) or meses < 1 or meses > 12:
            meses = 6
        
        # Calcular fecha de inicio (usando timezone.now() para awareness de zona horaria)
        fecha_fin = timezone.now()
        fecha_inicio = fecha_fin - relativedelta(months=meses)
        
        logger.debug(f"Obteniendo datos para {meses} meses desde {fecha_inicio.date()}")
        
        # Obtener datos de postulaciones por mes
        postulaciones_por_mes = list(
            Postulacion.objects
            .filter(fecha_postulacion__gte=fecha_inicio, fecha_postulacion__lte=fecha_fin)
            .annotate(mes=TruncMonth('fecha_postulacion'))
            .values('mes')
            .annotate(
                postulantes=Count('id'),
                graduados=Count('id', filter=Q(estado_general='TITULADO')),
                pendientes=Count('id', filter=Q(estado_general='EN_PROCESO')),
                aprobados=Count('id', filter=Q(estado_general='APROBADO')),
                rechazados=Count('id', filter=Q(estado_general='RECHAZADO'))
            )
            .order_by('mes')
        )
        
        # Obtener datos de documentos por mes
        documentos_por_mes = list(
            DocumentoPostulacion.objects
            .filter(fecha_subida__gte=fecha_inicio, fecha_subida__lte=fecha_fin)
            .annotate(mes=TruncMonth('fecha_subida'))
            .values('mes')
            .annotate(total=Count('id'))
            .order_by('mes')
        )
        
        # Mapear a diccionarios para búsqueda rápida
        postulaciones_dict = {item['mes']: item for item in postulaciones_por_mes}
        documentos_dict = {item['mes']: item for item in documentos_por_mes}
        
        # Generar lineChartData (Progreso General por Mes)
        lineChartData = []
        for i in range(meses):
            fecha = fecha_inicio + relativedelta(months=i)
            mes_label = fecha.strftime('%b')  # Ene, Feb, Mar, etc.
            
            # ✅ FIX: hacer las fechas compatible para búsqueda en diccionario
            # TruncMonth retorna datetime(year, month, 1, 0, 0, 0, tz)
            # Así que debemos buscar con esa exactitud
            fecha_key = fecha.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            postulacion = postulaciones_dict.get(fecha_key, {})
            
            lineChartData.append({
                'mes': mes_label,
                'graduados': int(postulacion.get('graduados', 0) or 0),
                'pendientes': int(postulacion.get('pendientes', 0) or 0),
                'aprobados': int(postulacion.get('aprobados', 0) or 0),
            })
        
        # Generar barChartData (Postulantes & Documentos por Semana/Mes)
        barChartData = []
        for i in range(meses):
            fecha = fecha_inicio + relativedelta(months=i)
            mes_label = f'Sem {i+1}' if meses == 6 else fecha.strftime('%b')
            
            # ✅ FIX: hacer las fechas compatible para búsqueda en diccionario
            fecha_key = fecha.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            postulacion = postulaciones_dict.get(fecha_key, {})
            documento = documentos_dict.get(fecha_key, {})
            
            barChartData.append({
                'semana': mes_label if meses == 6 else 'Mes ' + str(i+1),
                'postulantes': int(postulacion.get('postulantes', 0) or 0),
                'documentos': int(documento.get('total', 0) or 0),
            })
        
        # Generar pieChartData (Distribución de Estados)
        estados_stats = Postulacion.objects.values('estado_general').annotate(
            total=Count('id')
        ).order_by('-total')
        
        # Mapeo de colores
        estado_colors = {
            'TITULADO': '#10b981',       # Green - Completado
            'APROBADO': '#3b82f6',        # Blue - Por Revisar
            'EN_PROCESO': '#f59e0b',      # Amber - En Proceso
            'RECHAZADO': '#ef4444',       # Red - Rechazado
        }
        
        estado_nombres = {
            'TITULADO': 'Completado',
            'APROBADO': 'Por Revisar',
            'EN_PROCESO': 'En Proceso',
            'RECHAZADO': 'Rechazado',
        }
        
        pieChartData = []
        for estado_item in estados_stats:
            estado = estado_item.get('estado_general', 'DESCONOCIDO')
            total = int(estado_item.get('total', 0) or 0)
            
            if total > 0:  # Solo incluir si hay registros
                pieChartData.append({
                    'name': estado_nombres.get(estado, estado),
                    'value': total,
                    'color': estado_colors.get(estado, '#6b7280'),  # Gray default
                })
        
        # Si no hay datos, usar valores vacíos pero válidos
        if not lineChartData:
            lineChartData = [{'mes': f'Mes {i+1}', 'graduados': 0, 'pendientes': 0, 'aprobados': 0} for i in range(meses)]
        if not barChartData:
            barChartData = [{'semana': f'Sem {i+1}', 'postulantes': 0, 'documentos': 0} for i in range(meses)]
        if not pieChartData:
            # 100 en lugar de 1 para que el pie chart se vea al 100%
            pieChartData = [{'name': 'Sin datos', 'value': 100, 'color': '#d1d5db'}]
        
        response = {
            'lineChartData': lineChartData,
            'barChartData': barChartData,
            'pieChartData': pieChartData,
            'error': None
        }
        
        logger.debug(f"✅ Chart data generado: {len(lineChartData)} meses, {len(pieChartData)} estados")
        return response
        
    except Exception as e:
        logger.error(f"❌ Error en get_dashboard_chart_data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        # Retornar mock con indicador de error
        mock_response['error'] = str(e)
        return mock_response


def generar_pdf_dashboard(data: dict, fecha_inicio: str | None, fecha_fin: str | None, year: str | None) -> HttpResponse:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch/2, bottomMargin=inch/2)
    
    styles = getSampleStyleSheet()
    styles['h1'].textColor = colors.HexColor('#1A6FAB')
    styles['h2'].textColor = colors.HexColor('#1A6FAB')
    story = []

    # --- Logo y Título ---
    # NOTA: ReportLab tiene soporte limitado para SVG. Se recomienda usar un logo en formato PNG o JPG.
    # Coloca tu logo en 'static/logo.png' en la raíz del proyecto.
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=1.2*inch, height=0.6*inch)
        logo.hAlign = 'LEFT'
        story.append(logo)
        story.append(Spacer(1, 0.1*inch))

    story.append(Paragraph("Reporte de Dashboard Académico", styles['h1']))
    story.append(Spacer(1, 0.2*inch))

    # --- Rango de Fechas ---
    if year:
        story.append(Paragraph(f"Año Fiscal: {year}", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
    elif fecha_inicio or fecha_fin:
        rango = f"Periodo: {fecha_inicio or 'N/A'} al {fecha_fin or 'N/A'}"
        story.append(Paragraph(rango, styles['Normal']))
        story.append(Spacer(1, 0.2*inch))

    # --- Colores Corporativos ---
    primary_color = colors.HexColor('#1A6FAB')
    background_light = colors.HexColor('#EFF8FF')
    text_light = colors.white

    # --- Resumen General ---
    story.append(Paragraph("Resumen General", styles['h2']))
    cards_data = [
        ['Modalidades Activas', data.get('total_modalidades', 0)],
        ['Postulantes Registrados', data.get('total_postulantes', 0)],
        ['Postulaciones Totales', data.get('total_postulaciones', 0)],
        ['Documentos Cargados', data.get('total_documentos', 0)],
        ['Total Titulados', data.get('total_titulados', 0)],
        ['Promedio Proceso (días)', f"{data.get('tiempo_promedio_proceso_dias', 0.0):.2f}"],
    ]
    cards_table = Table(cards_data, colWidths=[3*inch, 1.5*inch])
    cards_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), background_light),
        ('GRID', (0, 0), (-1, -1), 1, primary_color),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, -1), primary_color),
    ]))
    story.append(cards_table)
    story.append(Spacer(1, 0.3*inch))

    # --- Distribución por Estado ---
    story.append(Paragraph("Distribución por Estado General", styles['h2']))
    estado_data = [['Estado', 'Total']]
    for item in data.get('postulaciones_por_estado_general', []):
        estado_data.append([item['estado'].replace('_', ' '), item['total']])
    
    estado_table = Table(estado_data, colWidths=[3*inch, 1.5*inch])
    estado_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), text_light),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, primary_color),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ]))
    story.append(estado_table)

    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte-dashboard.pdf"'
    return response


def estadisticas_tutores(year=None, carrera_id=None) -> list[dict]:
    """
    Calcula estadísticas de rendimiento de tutores:
    - Total de alumnos titulados.
    - Tiempo promedio de titulación (en días).
    """
    try:
        queryset = Postulacion.objects.filter(estado_general='TITULADO').exclude(tutor__isnull=True).exclude(tutor='')

        if carrera_id:
            try:
                carrera_value = str(carrera_id).strip()
                if carrera_value:
                    queryset = queryset.filter(postulante__carrera__iexact=carrera_value)
            except (ValueError, TypeError):
                pass

        if year:
            try:
                year_int = int(year)
                # Filtrar por año de titulación (estimado por la última revisión de documento)
                queryset = queryset.annotate(
                    fecha_fin_filter=Coalesce(Max('documentos__fecha_revision'), Now())
                ).filter(fecha_fin_filter__year=year_int)
            except (ValueError, TypeError):
                pass

        stats = (
            queryset.annotate(
                fecha_fin=Coalesce(Max('documentos__fecha_revision'), Now()),
                duracion=ExpressionWrapper(
                    F('fecha_fin') - F('fecha_postulacion'),
                    output_field=DurationField(),
                )
            )
            .values('tutor')
            .annotate(
                total_titulados=Count('id', distinct=True),
                tiempo_promedio=Avg('duracion')
            )
            .order_by('-total_titulados')
        )

        results = []
        for item in stats:
            try:
                # Validación segura de campos
                tutor_nombre = (item.get('tutor') or '').strip()
                total_titulados = int(item.get('total_titulados') or 0)
                tiempo_promedio = item.get('tiempo_promedio')
                
                # Cálculo seguro de días
                tiempo_dias = 0.0
                if tiempo_promedio is not None:
                    try:
                        tiempo_dias = round(tiempo_promedio.total_seconds() / 86400, 2)
                    except (AttributeError, TypeError):
                        tiempo_dias = 0.0
                
                results.append({
                    'tutor_id': _tutor_hash(tutor_nombre),
                    'nombre': tutor_nombre,
                    'total_titulados': total_titulados,
                    'tiempo_promedio_dias': tiempo_dias
                })
            except Exception as e:
                print(f"⚠️ Error procesando item de tutor: {e}")
                continue
        
        return results
    
    except Exception as e:
        import traceback
        print(f"❌ Error en estadisticas_tutores: {str(e)}")
        print(traceback.format_exc())
        return []


def detalle_alumnos_titulados_por_tutor(tutor_id: int) -> list[dict]:
    """
    Obtiene el detalle de los alumnos titulados bajo la tutoría de un docente específico.
    """
    try:
        tutor_id_value = int(tutor_id)
    except (ValueError, TypeError):
        return []

    try:
        queryset = Postulacion.objects.filter(
            estado_general='TITULADO',
        ).exclude(tutor__isnull=True).exclude(tutor='').select_related('postulante', 'modalidad').annotate(
            fecha_fin=Coalesce(Max('documentos__fecha_revision'), Now()),
            duracion=ExpressionWrapper(
                F('fecha_fin') - F('fecha_postulacion'),
                output_field=DurationField(),
            )
        ).order_by('-fecha_fin')

        results = []
        for p in queryset:
            try:
                if _tutor_hash((p.tutor or '').strip()) == tutor_id_value:
                    duracion_dias = 0
                    if hasattr(p, 'duracion') and p.duracion:
                        try:
                            duracion_dias = p.duracion.days
                        except (AttributeError, TypeError):
                            duracion_dias = 0
                    
                    fecha_fin = None
                    if hasattr(p, 'fecha_fin') and p.fecha_fin:
                        try:
                            fecha_fin = p.fecha_fin.strftime('%Y-%m-%d')
                        except (AttributeError, TypeError):
                            fecha_fin = None
                    
                    results.append({
                        'postulacion_id': p.id,
                        'alumno_nombre': f"{p.postulante.nombre} {p.postulante.apellido}",
                        'alumno_codigo': p.postulante.codigo_estudiante,
                        'modalidad': p.modalidad.nombre if p.modalidad else 'N/A',
                        'titulo_trabajo': p.titulo_trabajo,
                        'gestion': p.gestion,
                        'fecha_inicio': p.fecha_postulacion.strftime('%Y-%m-%d') if p.fecha_postulacion else 'N/A',
                        'fecha_fin': fecha_fin,
                        'duracion_dias': duracion_dias,
                    })
            except Exception as e:
                print(f"⚠️ Error procesando alumno: {e}")
                continue
        
        return results
    except Exception as e:
        import traceback
        print(f"❌ Error en detalle_alumnos_titulados_por_tutor: {str(e)}")
        print(traceback.format_exc())
        return []


def _tutor_hash(value: str) -> int:
    return zlib.adler32(value.encode('utf-8')) & 0xFFFFFFFF


def reporte_eficiencia_carreras(year=None) -> list[dict]:
    """
    Genera un reporte de eficiencia por carrera universitaria:
    - Total de procesos iniciados.
    - Tasa de titulación (Titulados / Iniciados).
    - Tiempo promedio de titulación (en días).
    """
    try:
        queryset = Postulacion.objects.select_related('postulante')

        if year:
            try:
                year_int = int(year)
                queryset = queryset.filter(fecha_postulacion__year=year_int)
            except (ValueError, TypeError):
                pass

        # Agrupar por carrera para conteos
        stats = queryset.values(
            'postulante__carrera',
        ).annotate(
            total_iniciados=Count('id'),
            total_titulados=Count('id', filter=Q(estado_general='TITULADO')),
        ).order_by('-total_titulados')

        # Calcular tiempos promedio solo para los titulados
        tiempos_queryset = queryset.filter(estado_general='TITULADO').annotate(
            fecha_fin=Coalesce(Max('documentos__fecha_revision'), Now()),
            duracion=ExpressionWrapper(
                F('fecha_fin') - F('fecha_postulacion'),
                output_field=DurationField(),
            )
        ).values('postulante__carrera').annotate(
            tiempo_promedio=Avg('duracion')
        )
        
        tiempos_map = {item['postulante__carrera']: item['tiempo_promedio'] for item in tiempos_queryset}

        results = []
        for item in stats:
            try:
                carrera = item['postulante__carrera'] or 'Sin Carrera Asignada'
                total_iniciados = int(item['total_iniciados'] or 0)
                total_titulados = int(item['total_titulados'] or 0)
                
                # Cálculo seguro de tasa de titulación
                tasa = 0.0
                if total_iniciados > 0:
                    tasa = round((total_titulados / total_iniciados * 100), 2)
                
                # Cálculo seguro de tiempo promedio
                tiempo_dias = 0.0
                tiempo_promedio = tiempos_map.get(item['postulante__carrera'])
                if tiempo_promedio is not None:
                    try:
                        tiempo_dias = round(tiempo_promedio.total_seconds() / 86400, 2)
                    except (AttributeError, TypeError):
                        tiempo_dias = 0.0
                
                results.append({
                    'carrera': carrera,
                    'total_iniciados': total_iniciados,
                    'total_titulados': total_titulados,
                    'tasa_titulacion': tasa,
                    'tiempo_promedio_dias': tiempo_dias
                })
            except Exception as e:
                print(f"⚠️ Error procesando carrera: {e}")
                continue
        
        return results
    except Exception as e:
        import traceback
        print(f"❌ Error en reporte_eficiencia_carreras: {str(e)}")
        print(traceback.format_exc())
        return []


def generar_excel_tutores(data: list[dict]) -> HttpResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendimiento Tutores"

    # Encabezados
    headers = ["ID", "Nombre", "Total Titulados", "Tiempo Promedio (Días)"]
    ws.append(headers)

    # Estilo de encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A6FAB", end_color="1A6FAB", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for item in data:
        ws.append([
            item['tutor_id'],
            item['nombre'],
            item['total_titulados'],
            item['tiempo_promedio_dias']
        ])

    # Ajustar ancho de columnas
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value) or "") for cell in column_cells)
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = max_length + 2

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estadisticas_tutores.xlsx"'
    return response
